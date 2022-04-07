#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
@version:  1.0
@author:   Jianzhang Zhang, <jianzhang.zhang@foxmail.com>
@file:     synonymyprocess.py
@time:     2016-10-06
@function: 常用的文件操作方法
"""

import os
import re
import csv
import xlsxwriter
from openpyxl import load_workbook
from .stringUtils import jiema,encodeUTF8
import dill

def readExcelToList(excelFile):
	'''
	读取excel文件到list中
	:param excelFile: excel文件路径
	:return table: 保存excel文件中数据的列表, 格式list(list)
	'''
	wb = load_workbook(excelFile)
	ws = wb.active
	numOfRow = ws.max_row
	numOfCol = ws.max_column
	table = []
	for row in range(1,numOfRow+1):
		rowList = []
		for col in range(1,numOfCol+1):
			cellValue = ws.cell(row=row, column=col).value
			rowList.append(cellValue)
		table.append(rowList)
	return table

def getFileList(dir, fileList):
	"""
  遍历一个目录,输出所有文件名
  param dir: 待遍历的文件夹
  param filrList : 保存文件名的列表
  return fileList: 文件名列表
  """
	if os.path.isfile(dir):
		fileList.append(dir)
	elif os.path.isdir(dir):
		for s in os.listdir(dir):
			# 如果需要忽略某些文件夹，使用以下代码
			# if s == "xxx":
			# continue
			newDir = os.path.join(dir, s)
			getFileList(newDir, fileList)
	return fileList


def readStrFromFile(filePath):
	"""
  从文件中读取字符串str
  param filePath: 文件路径
  return string : 文本字符串
  """
	with open(filePath, "rb") as f:
		string = jiema(f.read().strip())
	return string


def readLinesFromFile(filePath):
	"""
  从文件中读取字符串列表list
  param filePath: 文件路径
  return lines  : 文本字符串列表
  """
	
	with open(filePath, "rb") as f:
		content = jiema(f.read().strip())
	lines = content.split("\n")
	return lines


def writeStrToFile(filePath, string):
	"""
  将字符串写入文件中
  param filePath: 文件路径
  param string  : 字符串str
  """
	with open(filePath, "wb") as f:
		f.write(encodeUTF8(string))


def appendStrToFile(filePath, string):
	"""
  将字符串追加写入文件中
  param filePath: 文件路径
  param string  : 字符串str
  """
	with open(filePath, "ab") as f:
		f.write(encodeUTF8(string))


def dumpToFile(filePath, content):
	"""
  将数据类型序列化存入本地文件
  param filePath: 文件路径
  param content : 待保存的内容(list, dict, tuple, ...)
  """
	with open(filePath, "wb") as f:
		dill.dump(content, f)


def loadFromFile(filePath):
	"""
  从本地文件中加载序列化的内容
  param filePath: 文件路径
  return content: 序列化保存的内容(e.g. list, dict, tuple, ...)
  """
	with open(filePath,'rb') as f:
		content = dill.load(f)
	return content


def loadJson(filePath):
	'''
	从本地加载json文件
	:param filePath: 文件路径
	:return content: json格式的内容
	'''
	import json
	with open(filePath) as f:
		content = json.load(f)
	return content


def dumpJson(filePath,obj):
	'''
	将json内容写入本地文件
	:param content: json格式的内容
	:param filePath: 文件路径
	:return:
	'''
	import json
	content = json.dumps(obj, ensure_ascii=False, sort_keys=True, indent=4)
	writeStrToFile(filePath,content)

def appendToJson(filePath,obj):
	'''
	将json内容写入本地文件
	:param content: json格式的内容
	:param filePath: 文件路径
	:return:
	'''
	import json
	content = json.dumps(obj, ensure_ascii=False, sort_keys=True, indent=4)
	appendStrToFile(filePath,content)
	
def writeToExcel(excel, tableList):
	'''
	将list(list)内容写入excel文件
	:param excel: excel文件路径
	:param tableList: 待写入的内容, 格式为list(list)
	:return: 
	'''
	wb = xlsxwriter.Workbook(excel)
	ws = wb.add_worksheet('sheet1')
	for i, itemList in enumerate(tableList):
		newItemList = [jiema(item) if item else "" for item in itemList]
		for j, item in enumerate(newItemList):
			ws.write(i, j, item)
	wb.close()
	
def formatFilename(rawName):
	'''
	将文件名转化为不包含非法字符的文件名
	:param rawName: 原始文件名
	:return: 合法文件名
	'''
	return re.sub('[\\\\/:*?"<>|\s]', '_', rawName)

def appendToCsv(filePath, table):
	with open(filePath,"a") as f:
		writer = csv.writer(f)
		for row in table:
			# newRow = [encodeUTF8(item) if item else "" for item in row]
			# writer.writerow(newRow)
			writer.writerow(row)
	

def readCsvToList(csvFile):
	table = []
	with open(csvFile) as f:
		csvReader = csv.reader(f)
		for row in csvReader:
			table.append(row)
	return table

def writeToCsv(filePath,table):
	with open(filePath,"w") as f:
		writer = csv.writer(f)
		for row in table:
			# newRow = [encodeUTF8(item) if item else "" for item in row]
			# writer.writerow(newRow)
			writer.writerow(row)

def writeToTsv(filePath,table):
	newTable = []
	for line in table:
		newLine = []
		for item in line:
			if item.__class__ not in [str,unicode]:
				newLine.append(str(item))
			else:
				newLine.append(item)
		newTable.append("\t".join(newLine))
	writeStrToFile(filePath,"\n".join(newTable))

def readTsvToTable(tsvFile):
	table = []
	lines = readLinesFromFile(tsvFile)
	for line in lines:
		table.append(line.split("\t"))
	return table

def readLinesToTable(txtFile):
	table = []
	lines = readLinesFromFile(txtFile)
	for line in lines:
		table.append(line.split(" "))
	return table
	

def invertDict(mydict):
	inverted_dict = dict([[v, k] for k, v in mydict.items()])
	return inverted_dict
